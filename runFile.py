# -*- coding: utf-8 -*-
"""
Created on Sat Aug 19 23:03:16 2017

@author: Evyatar
"""


#######################################
####TYPE "%matplotlib tk" at command line
#######################################



from InstancePieceWiseLinear import *
import openpyxl 

num_fig=1



def show_instance(cons,eps):
    global num_fig,tup,cons_norm,cr
    fs=30
    cons_norm = cons.normalize()
    cr,q,T,cr_lp = cons_norm.approxOptStr(eps)
    detOptStr,crDetOptStr = cons_norm.optDetStr()
    #print detOptStr,crDetOptStr
    cons.plot(num_fig=num_fig)

    plt.title("Original instance (CR={0})".format(cr),fontsize=fs)
    plt.plot(np.array(T)*cons_norm.norm_time,cr+cr_lp,"k*-")
    plt.plot(np.array(T[:-1])*cons_norm.norm_time,np.cumsum(q)[:-1],"g*-")
    if detOptStr !=np.inf:
        plt.plot(detOptStr*cons_norm.norm_time,crDetOptStr,"dm",markersize=20)
    else:
        #print "inf",plt.xticks()[0][-1]
        plt.plot(plt.xlim()[1],crDetOptStr,"dm",markersize=20)
    plt.legend()
    plt.legend(["rent","buy","competitive ratio","strategy"],loc = 0)
    num_fig +=1
   
    cons_norm.plot(num_fig=num_fig)
    plt.title("Normalized instance (CR={0})".format(cr),fontsize=fs)
    plt.plot(T,cr+cr_lp,"k*-")
    plt.plot(T[:-1],np.cumsum(q)[:-1],"g*-")
    if detOptStr !=np.inf:
        plt.plot(detOptStr,crDetOptStr,"dm",markersize=20)
    else:
        #print "inf",plt.xticks()[0][-1]
        plt.plot(plt.xlim()[1],crDetOptStr,"dm",markersize=20)
        
    
    #eps,c,q_inf at graph or otherwise
    print "Switch cost: {0} ".format(c)
    print "Switch time of optimal determinstic startegy : {0}".format(detOptStr)  
    print "Competitive ratio of optimal determinstic startegy : {0}".format(crDetOptStr)  
    print "Approximaton factor: {0}".format(eps)   
    print "Competitive ratio of optimal probabiltstic strategy : {0}".format(cr)   
    print "Probability for start at \"buy\"  of optimal probabiltstic strategy : {0}".format(q[0])
    print "Probability for ending at \"rent\"  of optimal probabiltstic strategy : {0}".format(q[-1])
    
    
    
    return
   
   
   
   
"""
Example:
    
br0,bb0 = 0.7,1.0
ts=[0.0,1.0,1.3,2.2]
ar_s = [0.9,0.9,0.0,0.3,0]
ab_s= [0.0,0.0,0.0,0.2]

c=1
eps=0.05
"""

   

#checking -  exceptions,degenerate
#points
#tight recognition
#show eps,c,q_inf at graph or othersise
#manual


#TODO1 - send code - 3 days
#TODO2 - det,prob - 1.9.17
#TODO3 - single  GNUPLOT - 8.9.17
#TODO4 - intro - 1.10

br0,bb0 = 0.7,1.0
ts=[0.0,1.0,1.3,2.2,5.0]
ar_s = [0.9,0.9,1.5,1.5,2.0]
ab_s= [0.0,0.0,0.0,0.2,0.0]

c=1
eps=0.01




#cons = create_instance(br0,bb0,ar_s,ab_s,ts,c)    

#show_instance(cons,eps)

def get_input():
    strFlag="""Which way you prefer to insert the parameters to the instance:
1. Slopes
2. Points
"""
    mode = input(strFlag)
    if mode==1:
        
        br0=input("Insert the initial cost of the rent function\n")
        bb0=input("Insert the initial cost of the buy function\n")
        ts=input("Insert the time intervals edges\n0,")
        if ((type(ts)==int) or (type(ts)==float)):
            ts=[0,ts]
        else:
            ts=list(ts)
            ts.insert(0,0)
        ar_s = list(input("Insert the slopes of the rent function edges\n"))
        ab_s = list(input("Insert the slopes of the buy function edges\n"))
        c = input("Insert the switch cost\n")
        eps = input("Insert the approximantion factor\n")

#get_input()
        



def readPieceWiseLinear(ws,numRow):
    colStartCost =6
    indCol=colStartCost
    #print "read PWL"
    cost = ws.cell(row=numRow,column=indCol).value
    indCol +=1
    slope = ws.cell(row=numRow,column=indCol).value
    indCol +=1
    ts=[0]
    lins=[Linear(slope,cost)]

    
    while ws.cell(row=numRow,column=indCol).value is not None:  
        t=ws.cell(row=numRow,column=indCol).value
        ts.append(t)
        indCol +=1
        cost =lins[-1][t]

        slope = ws.cell(row=numRow,column=indCol).value
        indCol +=1
        #print t,slope,cost
        lins.append(Linear(slope,cost-slope*t))
        
       
    pwl = PieceWiseLinear(lins,ts)

        
       
        
    return pwl

"""
lins_r,lins_b=[Linear(ar_s[0],br0)],[Linear(ab_s[0],bb0)]
    for ind,t in enumerate(ts[1:]):
        yr,yb=lins_r[-1][t],lins_b[-1][t]
        lins_r.append(Linear(ar_s[ind+1],-ar_s[ind+1]*t+yr))
        lins_b.append(Linear(ab_s[ind+1],-ab_s[ind+1]*t+yb))
    pr=PieceWiseLinear(lins_r,ts)
    pb=PieceWiseLinear(lins_b,ts)
    return InstancePieceWiseLinear(pr,pb,c)
"""




def cellValid(cell):
    if cell.value is None:
        return False
    if cell.value.lower()=="v":
        return True
    return False
  



def mergePWL(pwl1,pwl2):
    def next_trio(pwl,ind):
        if ind<len(pwl)-1:
            return {"index":ind+1,"Linear":pwl.lins[ind+1],"time":pwl.times[ind+1]}
        else:
            return  {"index":len(pwl),"Linear":pwl.lins[-1],"time":np.inf}
            len(pwl),pwl.lins[-1],np.inf
    #We will merge the times without intersections
    ts=[0]
    lins1,lins2=[pwl1.lins[0]],[pwl2.lins[0]]
    ind1,ind2 = 0,0
    #next1,next2=pwl1.times[ind1],pwl2.times[ind2]
    #lin1,lin2 = pwl1.lins[ind1],pwl2.lins[ind2]
    #next1,next2=pwl1.times[ind1+1] if ind1==len(pwl1)-1 else np.inf,pwl2.times[ind2+1] if ind2==len(pwl2)-1 else np.inf
    #ind1,ind2 = min(ind1+1,len(pwl1)-1)  , min(ind2+1,len(pwl2)-1)    
    
    #we iterate on (index,linear,time)
    trio1 = next_trio(pwl1,ind1)
    trio2 = next_trio(pwl2,ind2)
    while (trio1["time"]<np.inf) or (trio2["time"]<np.inf):
        if trio1["time"]==trio2["time"]:
            ts.append( trio1["time"])
            lins1.append(trio1["Linear"])
            lins2.append(trio2["Linear"])
            trio1 = next_trio(pwl1,trio1["index"])
            trio2 = next_trio(pwl2,trio2["index"])
        elif trio1["time"]<trio2["time"]:
            ts.append( trio1["time"])
            lins1.append(trio1["Linear"])
            lins2.append(lins2[-1][:])
            trio1 = next_trio(pwl1,trio1["index"])
        else:
            ts.append( trio2["time"])
            lins2.append(trio2["Linear"])
            lins1.append(lins1[-1][:])
            trio2 = next_trio(pwl2,trio2["index"])
        
    return PieceWiseLinear(lins1,ts),PieceWiseLinear(lins2,ts)



def get_input_xl():
    file_excel ="Instances.xlsx"
    wb = openpyxl.load_workbook(filename = file_excel)
    ws = wb["Instances"]
    eps=0.01
    switchCost=np.inf
    rowStartInstances,colID,colValid,colEpsilon,colSwitchCost= 2,1,3,4,5
    indRow=rowStartInstances
    listPWL=[]
    while ((ws.cell(row=indRow,column=colID).value is not None) and (len(listPWL)<2)):  
        if cellValid(ws.cell(row=indRow,column=colValid)):
            eps = min(eps,ws.cell(row=indRow,column=colEpsilon)) if ws.cell(row=indRow,column=colEpsilon).value is not None else eps
            switchCost = min(switchCost,ws.cell(row=indRow,column=colSwitchCost).value) if ws.cell(row=indRow,column=colSwitchCost).value is not None else switchCost
            listPWL.append(readPieceWiseLinear(ws,indRow))
            indRow +=1
        else:
            indRow +=1
            continue
    #print switchCost
    pwl1,pwl2 = mergePWL(listPWL[0],listPWL[1])
    if pwl1(0)<pwl2(0):
        return InstancePieceWiseLinear(pwl1,pwl2,switchCost),eps
    else:
        return InstancePieceWiseLinear(pwl2,pwl1,switchCost),eps
        
    

    

cons,eps = get_input_xl()  
show_instance(cons,eps)

